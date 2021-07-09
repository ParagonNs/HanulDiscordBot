#Excel.py
#code base ( richardlee-kr 's SuperBot )
#https://github.com/richardlee-kr/SuperBot

from openpyxl import load_workbook, Workbook

Workbook_DB = load_workbook("BotDB.xlsx", data_only = True)
WB_Sheet_userDB = Workbook_DB.create_sheet("User Data", 0 )
WB_Sheet_CountryDB = Workbook_DB.create_sheet("Country Data", 1 )

def Load_file():
    print("Excel.py의 Load_file 실행중")
    print("엑셀 파일을 불러옵니다.")
    Workbook_DB = load_workbook("BotDB.xlsx", data_only = True)
    WB_Sheet_userDB = Workbook_DB.create_sheet("User Data", 0 )
    WB_Sheet_CountryDB = Workbook_DB.create_sheet("Country Data", 1 )
    
def Save_file():
    print("Excel.py의 Save_file 실행중")
    print("엑셀 파일을 저장합니다.")
    Workbook_DB.save("BotDB.xlsx")
    Workbook_DB.close()
           
def Check_user_num():
    print("Excel.py의 Check_user_num 실행중")
    Load_file()
    _user_count = 0
    for _i in range( 2, WB_Sheet_userDB.max_row + 1 ):
        if WB_Sheet_userDB.cell( _i, User_data_list[1] ).value != None:
            _user_count = _user_count + 1
    print("데이터베이스에 등록되어있는 유저 수는 {0}명입니다.".format( _user_count ))
    Save_file()
    return _user_count

def Check_user( _name, _id ):
    print("Excel.py의 Check_user 실행중")
    Load_file()
    print("유저 {0}<{1}>가 존재하는지 확인합니다.".format( _name, hex(_id)))
    _total_user_num = Check_user_num()
    print("탐색을 시작합니다.")
    for _i in range(2, 3 + userNum):
        _i_name_value = WB_Sheet_userDB.cell( _i, User_data_list[1] ).value
        _i_id_value = WB_Sheet_userDB.cell( _i, User_data_list[0] ).value
        print("{0}번째 정보 탐색, {0}번째 줄의 이름 : {1}".format(_i, _i_name_value ))
        print("{0} 과/와 {1} 이/가 일치 여부 : {2}".format( _name, _i_name_value, _name == _i_name_value ))
        print("{0}번째 줄의 유저 아이디 : {1}".format(_i, _i_id_value ))
        print("{0} 과/와 {1} 이/가 일치 여부 : {2}".format( hex(_id), _i_id_value, hex(_id) == _i_id_value ))
        if _i_name_value == _name and _i_id_value == hex(_id):
            print("{0}번째 정보에서 유저 {1}<{2}>를 발견했습니다.".format( _i, _name, hex(_id)))
            print("유저 정보가 등록된 위치는 {0}번째 줄입니다.".format( _i ))
            Save_file()
            return True, _i
            break
        else:
            print("재탐색중...")
    Save_file()
    print("유저 {0}<{1}>을/를 발견하지 못했습니다.".format( _name, hex(_id)))
    return False, None
    
def Last_row_count():
    print("Excel.py의 Last_row_count 실행중")
    Load_file()
    for _i in range(2, WB_Sheet_userDB.max_row + 1):
        if WB_Sheet_userDB.cell( _i, 1 ).value is None:
            print("가장 처음으로 발견된 빈 행은 {{0}}행입니다.".format( _i ))
            Save_file()
	    return _i
            break
    
def Signup( _name, _id ):
    print("Excel.py의 Signup 실행중")
    Load_file()
    _row = Last_row_count()
    print("{{0}}행에 신규 유저 정보를 생성합니다.".format(_row))
    
    #Data list
    User_data_list[ hex(_id), _name, 100, 1, 0, 0, 0, 0, 1, 0, 0, 0, 500, 1, 0, 0 ]
    #id_0, name_1, credit_2, user main lv_3
    #resource A_4, resource B_5, resource C_6
    #resource D_7, user mining lv_8, user mining count lv_9
    #user mining amount lv_10, user mining luck lv_11, vault credit_12
    #user bank lv_13, total bank tax_14, total tax_15
    
    for _i in range( 0, len( User_data_list )):
        WB_Sheet_userDB.cell( _row, column = _i + 1, value = User_data_list[_i])
    
    Save_file()
    print("신규 유저 {{0}}님의 정보 생성이 완료되었습니다.".format(_name))
    
def Low_user_data( _name, _id ):
    print("Excel.py의 Low_user_data 실행중")
    Load_file()
    _checked_name, _checked_at = Check_user( _name, _id )
    if _checked_name
	_low_user_data = []
	for _i in range( 0, len( User_data_list )):
	    _low_user_data += WB_Sheet_userDB.cell( _checked_at, _i ).value
	Save_file()
	print("유저 {{}}님의 정보가 성공적으로 추출되었습니다.".format(_name))
        return _low_user_data
    else:
	Save_file()
	print("유저 {{}}님의 정보가 추출되지 않았습니다.".format(_name))
	return None
