#user.py
#code base ( richardlee-kr 's SuperBot )
#https://github.com/richardlee-kr/SuperBot

from openpyxl import load_workbook, Workbook

Workbook_DB = load_workbook("BotDB.xlsx", data_only = True)
WB_Sheet_userDB = Workbook_DB.create_sheet("User Data", 0 )
WB_Sheet_CountryDB = Workbook_DB.create_sheet("Country Data", 1 )

def Load_file():
    print("엑셀 파일을 불러옵니다.")
    Workbook_DB = load_workbook("BotDB.xlsx", data_only = True)
    WB_Sheet_userDB = Workbook_DB.create_sheet("User Data", 0 )
    WB_Sheet_CountryDB = Workbook_DB.create_sheet("Country Data", 1 )
    
def Save_file():
    print("엑셀 파일을 저장합니다.")
    Workbook_DB.save("BotDB.xlsx")
    Workbook_DB.close()
    
def checkUserNum():
    print("user.py의 checkUserNum 실행중")
    loadFile()
    count = 0
    for row in range(2, wus.max_row+1):
        if wus.cell(row,User_data_list[1]).value != None:
            count = count+1
        else:
            count = count
    return count

def checkFirstRow():
    print("user.py의 checkFirstRow 실행중")
    loadFile()
    for row in range(2, wus.max_row + 1):
        if wus.cell(row,1).value is None:
            return row
            break
    _result = wus.max_row+1
    saveFile()
    return _result

def checkUser(_name, _id):
    print("user.py의 checkUser 실핼중")
    print(str(_name) + "<" + str(_id) + ">의 존재 여부 확인")
    loadFile()
    userNum = checkUserNum()
    print("등록된 유저수: ", userNum)
    print("이름과 고유번호 탐색")
    print("")
    for row in range(2, 3+userNum):
        print(row, "번째 줄 name: ", wus.cell(row,User_data_list[1]).value)
        print("입력된 name: ", _name)
        print("이름과 일치 여부: ", wus.cell(row, User_data_list[1]).value == _name)
        print(row,"번째 줄 id: ", wus.cell(row, User_data_list[0]).value)
        print("입력된 id: ", hex(_id))
        print("유저 고유 번호 와 일치 여부: ", wus.cell(row, User_data_list[0]).value == hex(_id))
        if wus.cell(row, User_data_list[1]).value == _name and wus.cell(row, User_data_list[0]).value == hex(_id):
            print("등록된  이름과 유저 고유 번호를 발견")
            print("등록된  값의 위치: ",  row, "번째 줄")
            saveFile()
            return True, row
            break
        else:
            print("등록된 정보를 탐색 실패, 재탐색 실시")
    saveFile()
    print("발견 실패")
    return False, None
           
def Check_user_num():
    print("user.py의 Check_user_num 실행중")
    Load_file()
    _user_count = 0
    for _i in range( 2, WB_Sheet_userDB.max_row + 1 ):
        if WB_Sheet_userDB.cell( _i, User_data_list[1] ).value != None:
            _user_count = _user_count + 1
    print("데이터베이스에 등록되어있는 유저 수는 {0}명입니다.".format( _user_count ))
    Save_file()
    return _user_count

def Check_user( _name, _id ):
    print("user.py의 Check_user 실행중")
    Load_file()
    print("유저 {0}<{1}>가 존재하는지 확인합니다.".format( _name, hex(_id)))
    _total_user_num = Check_user_num()
    print("탐색을 시작합니다.")
    for _i in range(2, 3+userNum):
        _i_name_value = WB_Sheet_userDB.cell( _i, User_data_list[1] ).value
        _i_id_value = WB_Sheet_userDB.cell( _i, User_data_list[0] ).value
        print("{0}번째 정보 탐색, {0}번째 줄의 이름 : {1}".format(_i, _i_name_value ))
        print("{0} 과/와 {1} 이/가 일치 여부 : {2}".format( _name, _i_name_value, _name == _i_name_value ))
        print("{0}번째 줄의 유저 아이디 : {1}".format(_i, _i_id_value ))
        print("{0} 과/와 {1} 이/가 일치 여부 : {2}".format( hex(_id), _i_id_value, hex(_id) == _i_id_value ))
        if _i_name_value == _name and _i_id_value == hex(_id):
            print("{0}번째 정보에서 유저 {1}<{2}>를 발견했습니다.".format( _i, _name, hex(_id)))
            print("유저 정보가 등록된 위치는 {0}번째 줄입니다.".format( _i ))
            Save_file()
            return True, _i
            break
        else:
            print("재탐색중...")
    Save_file()
    print("유저 {0}<{1}>을/를 발견하지 못했습니다.".format( _name, hex(_id)))
    return False, None
    
def Signup( _name, _id ):
    print("user.py의 Signup 실행중")
    Load_file()
    _row = Last_row_count()
    print("가장 처음으로 발견된 빈 행은 {{0}}행입니다.".format(_row))
    print("{{0}}행에 신규 유저 정보를 생성합니다.".format(_row))
    
    #Data list
    User_data_list[ hex(_id), _name, 100, 1, 0, 0, 0, 0, 1, 0, 0, 0, 500, 1, 0, 0 ]
    #id, name, credit, 유저의 레벨( 메인 ), 유저가 가진 자원 A, 유저가 가진 자원 B
    #유저가 가진 자원 C, 유저가 가진 자원 D, 유저의 채굴 레벨, 채굴 횟수 업그레이드, 채굴량 업그레이드, 행운( 채굴 ) 업그레이드
    #은행에 보관된 유저의 크레딧, 유저의 은행 등급, 유저가 은행에 지불한 모든 수수료의 총합, 유저가 지불한 모든 세금의 총합
    
    for _i in range( 0, len( User_data_list )):
        WB_Sheet_userDB.cell( _row, column = _i + 1, value = User_data_list[_i])
    
    Save_file()
    print("신규 유저 {{0}}님의 정보 생성이 완료되었습니다.".format(_name))
